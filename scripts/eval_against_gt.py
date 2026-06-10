"""Measure the cockpit's auto-assembled records against the ground-truth sheet.

Usage:  python scripts/eval_against_gt.py P1 [NF,AF,SL,FL]

This closes the loop on "is the pipeline reliable": it runs the same geometry
pipeline the cockpit uses (locate cache → parse_block → make_pairings, control
reused) and compares the resulting numbers to the manually-extracted values, cell
by cell. Matching is value-anchored (a predicted treatment matches a GT row when
its mean_t equals that row's Xe), so it needs no AF→"artificial forest" semantic
map and can't be fooled by label wording.

Reports, per study: record recall (GT control-vs-treatment points recovered) and
per-field accuracy on mean_c / sd_c / mean_t / sd_t within a relative tolerance.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

from _gt import GT, ROOT

VALUE_FIELDS = ["mean_c", "sd_c", "mean_t", "sd_t", "n_c", "n_t"]
GT_OF = {"mean_c": "Xc", "sd_c": "Sc", "mean_t": "Xe", "sd_t": "Se", "n_c": "Nc", "n_t": "Ne"}


def _num(s):
    if s is None:
        return None
    s = str(s).replace("−", "-")
    m = re.search(r"[-+]?\d*\.?\d+", re.sub(r"[^0-9eE.+\-]", "", s))
    try:
        return float(m.group()) if m else None
    except (ValueError, AttributeError):
        return None


def _toks(s):
    return set(re.findall(r"[a-z0-9]+", str(s).lower()))


def gt_records(study):
    import openpyxl

    ws = openpyxl.load_workbook(GT, read_only=True, data_only=True)["Sheet1"]
    r3 = [c.value for c in ws[3]]
    varcols = [(i, re.sub(r"\s+", " ", str(r3[i])).strip()) for i in range(22, 491, 7) if r3[i] not in (None, "")]
    rows = [ri for ri in range(5, 77) if ws.cell(ri, 1).value == study]
    out = []
    for i, name in varcols:
        for ri in rows:
            xc = ws.cell(ri, i + 1).value
            if xc in (None, ""):
                continue
            out.append({
                "variable": name,
                "new_use": ws.cell(ri, 13).value,
                "Xc": _num(xc), "Sc": _num(ws.cell(ri, i + 2).value),
                "Xe": _num(ws.cell(ri, i + 3).value), "Se": _num(ws.cell(ri, i + 4).value),
                "Nc": _num(ws.cell(ri, i + 5).value), "Ne": _num(ws.cell(ri, i + 6).value),
            })
    return out


def predicted_records(study, labels):
    from metaextract import cockpit_cache
    from metaextract.sampling import find_sample_size_candidates
    from metaextract.tabular import make_pairings, parse_block

    doc, locout = cockpit_cache.load(ROOT, study)
    # n the cockpit would default to: first deterministic candidate (human-confirmed)
    n_cands = find_sample_size_candidates(doc)
    n = _num(n_cands[0].value) if n_cands else None
    preds = []
    for region in locout.regions:
        if region.citation.kind == "figure":
            continue
        block = doc.block(region.citation.block_id)
        rows = parse_block(block, labels)
        if not rows:
            continue
        chosen = rows[0] if len(rows) == 1 else max(
            rows, key=lambda r: len(_toks(r.label) & _toks(region.variable_name))
        )
        for rec in make_pairings(chosen.cells, labels, labels[0]):
            preds.append({
                "variable": region.variable_name,
                "mean_c": _num(rec["mean_c"]), "sd_c": _num(rec["sd_c"]),
                "mean_t": _num(rec["mean_t"]), "sd_t": _num(rec["sd_t"]),
                "n_c": n, "n_t": n,
            })
    return preds


def _close(a, b, tol=0.02):
    if a is None or b is None:
        return False
    if b == 0:
        return abs(a) < 1e-9
    return abs(a - b) / abs(b) <= tol


def main():
    study = sys.argv[1] if len(sys.argv) > 1 else "P1"
    labels = (sys.argv[2] if len(sys.argv) > 2 else "NF,AF,SL,FL").split(",")

    gt = gt_records(study)
    preds = predicted_records(study, labels)
    print(f"{study}: {len(gt)} GT control-vs-treatment points, {len(preds)} predicted (control={labels[0]})\n")

    # 1:1 variable alignment: each GT variable -> the predicted variable name with
    # the most token overlap (avoids cross-variable contamination in matching).
    pred_vars = sorted({p["variable"] for p in preds})
    var_map = {}
    for gv in {g["variable"] for g in gt}:
        best = max(pred_vars, key=lambda pv: len(_toks(pv) & _toks(gv)), default=None)
        if best and _toks(best) & _toks(gv):
            var_map[gv] = best

    field_tot = {f: 0 for f in VALUE_FIELDS}
    field_ok = {f: 0 for f in VALUE_FIELDS}
    matched = 0
    misses = []
    for g in gt:
        # within the aligned variable, pick the treatment whose mean_t is NEAREST Xe
        cands = [p for p in preds if p["variable"] == var_map.get(g["variable"])]
        cands = [p for p in cands if _close(p["mean_t"], g["Xe"])]
        p = min(cands, key=lambda p: abs(p["mean_t"] - g["Xe"]), default=None)
        if p is None:
            misses.append(g)
            continue
        matched += 1
        for f in VALUE_FIELDS:
            gv = g[GT_OF[f]]
            if gv is None:
                continue
            field_tot[f] += 1
            if _close(p[f], gv):
                field_ok[f] += 1

    print(f"=== record recall: {matched}/{len(gt)} = {matched / len(gt):.0%} ===")
    print("=== per-field accuracy (matched records, within 2%) ===")
    for f in VALUE_FIELDS:
        t = field_tot[f]
        acc = f"{field_ok[f] / t:.0%}" if t else "n/a"
        print(f"  {f:8} {field_ok[f]}/{t}  = {acc}")
    if misses:
        print(f"\n=== {len(misses)} GT points not matched (value-anchored) ===")
        for g in misses[:12]:
            print(f"  {g['variable'][:34]:34} new_use={g['new_use']!r:18} Xc={g['Xc']} Xe={g['Xe']}")


if __name__ == "__main__":
    main()
