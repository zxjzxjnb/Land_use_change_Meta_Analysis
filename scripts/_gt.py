"""Shared helpers for the eval/prepare scripts: .env loading + TaskSpec from GT."""

from __future__ import annotations

import os
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
GT = ROOT.parent / (
    "Data collection-ZA_soil_C_fractions_landuse_Change_Jan2025_"
    "original_improved format_Final.xlsx"
)
sys.path.insert(0, str(ROOT / "src"))


def load_env() -> None:
    env = ROOT / ".env"
    if env.exists():
        for line in env.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())


def _clean(s) -> str:
    return re.sub(r"\s+", " ", str(s)).strip()


def gt_spec_and_truth(study_id: str):
    """Return (all_target_vars, moderators, vars_present_for_study, gt_rows)."""
    import openpyxl

    ws = openpyxl.load_workbook(GT, read_only=True, data_only=True)["Sheet1"]
    r3 = [c.value for c in ws[3]]
    varcols = [(i, _clean(r3[i])) for i in range(22, 491, 7) if r3[i] not in (None, "")]
    # paper/site-level constant moderators (Old/New use are the pairing, not moderators)
    mods = [_clean(ws.cell(4, i + 1).value) for i in (4, 9, 10, 13, 14, 15, 19, 20, 21)]
    rows = [ri for ri in range(5, 77) if ws.cell(ri, 1).value == study_id]
    present = [
        name
        for i, name in varcols
        if any(ws.cell(ri, i + 1).value not in (None, "") for ri in rows)
    ]
    return [name for _, name in varcols], mods, present, rows


def build_spec(study_id: str):
    from metaextract.locate import TaskSpec

    targets, mods, present, rows = gt_spec_and_truth(study_id)
    spec = TaskSpec(domain="soil_C_fractions_landuse", target_variables=targets, moderators=mods)
    return spec, present, rows
