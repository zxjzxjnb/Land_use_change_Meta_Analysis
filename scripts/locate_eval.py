"""Run screen+locate on one source paper and score it against the GT sheet.

Usage:  python scripts/locate_eval.py P2  [model]

Loads .env manually (the package does not), builds a TaskSpec from the real
67-variable vocabulary in the ground-truth workbook, ingests the paper, runs
locate(), then reports screening recall, per-variable location with the actual
block text (so precision can be eye-checked), and any hallucinated block_ids.
"""

from __future__ import annotations

import os
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
GT = ROOT.parent / (
    "Data collection-ZA_soil_C_fractions_landuse_Change_Jan2025_"
    "original_improved format_Final.xlsx"
)


def load_env() -> None:
    env = ROOT / ".env"
    if env.exists():
        for line in env.read_text().splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())


def _clean(s) -> str:
    return re.sub(r"\s+", " ", str(s)).strip()


def gt_spec_and_truth(study_id: str):
    import openpyxl

    ws = openpyxl.load_workbook(GT, read_only=True, data_only=True)["Sheet1"]
    r3 = [c.value for c in ws[3]]
    varcols = [(i, _clean(r3[i])) for i in range(22, 491, 7) if r3[i] not in (None, "")]
    mods = [_clean(ws.cell(4, i + 1).value) for i in (9, 10, 11, 12, 13, 14, 15, 19, 20, 21)]
    rows = [ri for ri in range(5, 77) if ws.cell(ri, 1).value == study_id]
    present = [
        name
        for i, name in varcols
        if any(ws.cell(ri, i + 1).value not in (None, "") for ri in rows)
    ]
    return [name for _, name in varcols], mods, present, rows


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", s.lower()).strip()


def main() -> None:
    study = sys.argv[1] if len(sys.argv) > 1 else "P1"
    model = sys.argv[2] if len(sys.argv) > 2 else "gemini-2.5-flash"
    load_env()

    sys.path.insert(0, str(ROOT / "src"))
    from metaextract.ingest import ingest_pdf
    from metaextract.locate import TaskSpec, locate

    targets, mods, gt_present, gt_rows = gt_spec_and_truth(study)
    spec = TaskSpec(domain="soil_C_fractions_landuse", target_variables=targets, moderators=mods)

    pdf = ROOT / "data" / "sample_papers" / f"{study}.pdf"
    doc = ingest_pdf(pdf)
    print(f"{study}: {doc.n_pages} pages, {len(doc.blocks)} blocks | GT rows={gt_rows}")
    print(f"GT reports {len(gt_present)} variables | calling {model} ...\n")

    out = locate(doc, spec=spec, model=model)

    located_names = {_norm(r.variable_name): r for r in out.regions}
    by_block = {}
    for r in out.regions:
        by_block.setdefault(r.citation.block_id, []).append(r.variable_name)

    found, missed = [], []
    for v in gt_present:
        (found if _norm(v) in located_names else missed).append(v)

    print(f"=== LOCATION RECALL: {len(found)}/{len(gt_present)} GT variables located ===")
    if missed:
        print("MISSED:", missed)
    print()
    print("=== located block -> variables + actual text (eyecheck precision) ===")
    for bid, names in sorted(by_block.items()):
        blk = doc.block(bid)
        print(f"  [{bid}] p{blk.page}: {', '.join(names)}")
        print(f"        text: {blk.text[:120]!r}")
    extra = [r.variable_name for r in out.regions if _norm(r.variable_name) not in {_norm(v) for v in gt_present}]
    if extra:
        print("\n=== located but NOT in GT (possible false positives) ===")
        print(" ", extra)
    if out.problems:
        print("\n=== HALLUCINATED block_ids (C1 guard dropped) ===")
        for p in out.problems:
            print("  -", p)


if __name__ == "__main__":
    main()
